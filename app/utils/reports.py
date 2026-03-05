from collections import defaultdict
from datetime import datetime
from decimal import Decimal
from io import BytesIO
from typing import Dict, List, Optional, Tuple, Union

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.worksheet import Worksheet
from pandas import DataFrame

from app.db import DatabaseManager

Tasks = List[Dict[str, Union[str, Decimal]]]
Data = List[List[Union[str, Decimal]]]

db_manager: DatabaseManager = DatabaseManager()


def period_contains_2025(start_date: datetime, end_date: datetime) -> bool:
    """
    Checks if the given period contains any date in the year 2025.

    Args:
        start_date (datetime, optional): The start of the period.
        end_date (datetime, optional): The end of the period.

    Returns:
        bool: True if the period includes any date in 2025, False otherwise.
    """
    lower_bound: datetime = datetime(2024, 12, 31)
    upper_bound: datetime = datetime(2026, 1, 1)

    if not start_date and not end_date:
        return True
    elif not start_date and end_date and end_date > lower_bound:
        return True
    elif start_date and end_date and lower_bound < start_date < upper_bound and end_date >= start_date:
        return True
    elif start_date and lower_bound < start_date < upper_bound and not end_date:
        return True
    return False


def get_tasks_data(tasks: Tasks) -> Data:
    """
    Converts tasks object into list of lists for report generation.

    Args:
        tasks (Tasks): List of task records, each containing employee details,
            order information, and work metrics.

    Returns:
        tasks_data (Data): List of lists, where each inner list contains the data for one specific task.
    """

    employee_categories: Dict[str, str] = {
        "worker": "Рабочий",
        "specialist": "Специалист",
        "manager": "Руководитель",
    }

    tasks_data: Data = [
        [
            task["employee_name"],
            task["personnel_number"],
            employee_categories[task["employee_category"]],
            task["department"],
            task["order_number"],
            task["order_name"],
            task["work_name"],
            task["hours"],
            task["operation_date"],
        ]
        for task in tasks
    ]
    return tasks_data


def get_employees_data(tasks: Tasks) -> Data:
    """
    Groups tasks by employee and date, aggregating hours for report generation.

    Creates unique key from employee details and date to aggregate hours,
    then formats the result with translated category names.

    Args:
        tasks (Tasks): List of task records, each containing employee details,
            order information, and work metrics.

    Returns:
        employees_data (Data):  List of lists, where each inner list contains the data for one specific employee
            on specific date.
    """

    spent_hours_per_employee: Dict[Tuple[str, ...], Decimal] = defaultdict(Decimal)

    for task in tasks:
        key: Tuple[str, ...] = (
            task["employee_name"],
            task["personnel_number"],
            task["employee_category"],
            task["department"],
            task["operation_date"],
        )
        spent_hours_per_employee[key] += task["hours"]

    employee_categories: Dict[str, str] = {
        "worker": "Рабочий",
        "specialist": "Специалист",
        "manager": "Руководитель",
    }

    employees_data: Data = [
        [
            employee_details[0],
            employee_details[1],
            employee_categories[employee_details[2]],
            employee_details[3],
            operation_date,
            spent_hours,
        ]
        for (*employee_details, operation_date), spent_hours in spent_hours_per_employee.items()
    ]
    return employees_data


def get_basic_orders_data(tasks: Tasks, start_date: datetime, end_date: datetime) -> Data:
    """
    Returns orders data including planned, spent, and remaining hours.

    This function calculates the total spent hours per order based on the provided tasks,
    optionally includes spent hours for the year 2025, if the specified date range requires it.

    Args:
        tasks (Tasks): List of task records, each containing employee details,
            order information, and work metrics.
        start_date (datetime): The start date for selecting tasks from the database.
        end_date (datetime): The end date for selecting tasks from the database.

    Returns:
        orders_data (Data): List of lists, where each inner list contains the data for one specific order,
            including its number, name, planned hours, spent hours, and remaining hours.
            The final inner list in the outer list contains the totals for planned hours, spent hours,
            and remaining hours calculated across all orders in the dataset.
    """

    spent_hours_per_order: Dict[str, Decimal] = defaultdict(Decimal)

    for task in tasks:
        spent_hours_per_order[task["order_number"]] += task["hours"]

    if period_contains_2025(start_date=start_date, end_date=end_date):
        spent_hours_for_2025: Dict[str, Decimal] = db_manager.orders.get_spent_hours_for_2025()

        for order_number, spent_hours in spent_hours_for_2025.items():
            spent_hours_per_order[order_number] += spent_hours

    orders_data: Data = []

    order_numbers: Tuple[str] = tuple(spent_hours_per_order.keys())

    if order_numbers:
        planned_hours_per_order: List = db_manager.orders.get_planned_hours_per_order(order_numbers=order_numbers)

        for order_number, order_name, planned_hours in planned_hours_per_order:
            spent_hours: Decimal = spent_hours_per_order[order_number]
            remaining_hours: Decimal = planned_hours - spent_hours
            orders_data.append(
                [
                    order_number,
                    order_name,
                    planned_hours,
                    spent_hours,
                    remaining_hours,
                ]
            )

    planned_hours, spent_hours, remaining_hours = Decimal(0), Decimal(0), Decimal(0)

    for order_data in orders_data:
        planned_hours += order_data[2]
        spent_hours += order_data[3]
        remaining_hours += order_data[4]

    orders_data.append(["Итого", "", planned_hours, spent_hours, remaining_hours])
    return orders_data


def get_detailed_orders_data(tasks: Tasks) -> Data:
    """
    Returns order data with detailed information by types of work.

    This function extends the basic orders data by providing information on work types
    associated with each order. For each order it displays list of work types with their planned,
    spent and remaining hours.

    The function calculates spent hours of work types from the provided tasks for each order and
    displays all associated work types with their corresponding planned, spent and remaining hours.

    This provides detailed view of hour distribution across different work types within each order.

    Args:
        tasks (Tasks): List of task records, each containing employee details,
            order information, and work metrics.

    Returns:
        orders_data (Data): List of lists, where each inner list contains the data for one specific order,
            including its number, name, work name, planned hours, spent hours, and remaining hours.
    """

    spent_hours_per_work: Dict[str, Decimal] = defaultdict(Decimal)

    for task in tasks:
        key: Tuple[str, str] = (
            task["order_number"],
            task["work_name"],
        )
        spent_hours_per_work[key] += task["hours"]

    order_numbers, work_names = [], []

    for order_number, work_name in spent_hours_per_work.keys():
        order_numbers.append(order_number)
        work_names.append(work_name)

    orders_data: Data = []

    if order_numbers and work_names:
        planned_hours_per_work: List = db_manager.works.get_planned_hours_per_work(
            order_numbers=order_numbers,
            work_names=work_names,
        )

        for order_number, order_name, work_name, planned_hours in planned_hours_per_work:
            spent_hours: Decimal = spent_hours_per_work[(order_number, work_name)]
            remaining_hours: Decimal = planned_hours - spent_hours
            orders_data.append(
                [
                    order_number,
                    order_name,
                    work_name,
                    planned_hours,
                    spent_hours,
                    remaining_hours,
                ]
            )
        return orders_data


def configure_worksheet_columns(
    worksheet: Worksheet,
    column_widths: Optional[Dict[str, int]] = None,
    style_columns: Optional[List[str]] = None,
    bold_columns: Optional[List[str]] = None,
    merge_columns: Optional[List[str]] = None,
) -> None:
    border_style: Border = Border(
        Side(border_style="thin"),
        Side(border_style="thin"),
        Side(border_style="thin"),
        Side(border_style="thin"),
    )

    if column_widths:
        for column in worksheet.columns:
            column_letter: str = column[0].column_letter
            worksheet.column_dimensions[column_letter].width = column_widths[column_letter]

            if style_columns:
                for cell in column:
                    if cell.row > 1 and column_letter in style_columns:
                        cell.number_format = "0.00"
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    cell.border = border_style

    last_row: int = worksheet.max_row

    if bold_columns:
        for cell in worksheet[last_row]:
            if cell.column_letter in bold_columns:
                cell.font = Font(bold=True)

    if merge_columns:
        worksheet.merge_cells(f"{merge_columns[0]}{last_row}:{merge_columns[-1]}{last_row}")


def write_data_to_worksheet(
    workbook: Workbook,
    sheet_name: str,
    headers: List[str],
    data: Data,
    column_widths: Optional[Dict[str, int]] = None,
    style_columns: Optional[List[str]] = None,
    filter_columns: Optional[List[str]] = None,
    bold_columns: Optional[List[str]] = None,
    merge_columns: Optional[List[str]] = None,
) -> None:
    dataframe: DataFrame = DataFrame(data=data, columns=headers)
    worksheet: Worksheet = workbook.create_sheet()

    if sheet_name:
        worksheet.title = sheet_name

    for row in dataframe_to_rows(dataframe, index=False, header=True):
        worksheet.append(row)

    filter_range: str = worksheet.dimensions

    if filter_columns:
        filter_range: str = f"{filter_columns[0]}1:{filter_columns[-1]}1"

    worksheet.auto_filter.ref = filter_range

    for cell in worksheet[1]:
        cell.font = Font(bold=True)

    configure_worksheet_columns(
        worksheet=worksheet,
        column_widths=column_widths,
        style_columns=style_columns,
        bold_columns=bold_columns,
        merge_columns=merge_columns,
    )


def get_report_file(
    tasks_data: Data = [],
    employees_data: Data = [],
    basic_orders_data: Data = [],
    detailed_orders_data: Data = [],
) -> BytesIO:
    workbook: Workbook = Workbook()

    workbook.remove(workbook.active)

    write_data_to_worksheet(
        workbook=workbook,
        sheet_name="Назначенные задания",
        headers=[
            "ФИО сотрудника",
            "Таб. номер",
            "Категория сотрудника",
            "Наименование подразделения",
            "Номер заказа",
            "Наименование заказа",
            "Наименование работы",
            "Дата выполнения",
            "Затраченное время, ч",
        ],
        data=tasks_data,
        column_widths={"A": 28, "B": 18, "C": 18, "D": 22, "E": 22, "F": 44, "G": 44, "H": 24, "I": 18},
        style_columns=["I"],
        filter_columns=["A", "B", "C", "D", "E", "F", "G", "H"],
    )

    write_data_to_worksheet(
        workbook=workbook,
        sheet_name="Табель рабочего времени",
        headers=[
            "ФИО сотрудника",
            "Таб. номер",
            "Категория сотрудника",
            "Наименование подразделения",
            "Дата выполнения",
            "Затраченное время, ч",
        ],
        data=employees_data,
        column_widths={"A": 28, "B": 18, "C": 18, "D": 22, "E": 24, "F": 18},
        style_columns=["F"],
        filter_columns=["A", "B", "C", "D", "E"],
    )

    write_data_to_worksheet(
        workbook=workbook,
        sheet_name="Сводка по заказам",
        headers=[
            "Номер заказа",
            "Наименование заказа",
            "Плановая трудоемкость, ч",
            "Фактическая трудоемкость, ч",
            "Остаточная трудоемкость, ч",
        ],
        data=basic_orders_data,
        column_widths={"A": 22, "B": 44, "C": 22, "D": 22, "E": 22},
        style_columns=["C", "D", "E"],
        filter_columns=["A", "B"],
        bold_columns=["A", "B", "C", "D", "E"],
        merge_columns=["A", "B"],
    )

    write_data_to_worksheet(
        workbook=workbook,
        sheet_name="Детализация по заказам",
        headers=[
            "Номер заказа",
            "Наименование заказа",
            "Наименование работы",
            "Плановая трудоемкость, ч",
            "Фактическая трудоемкость, ч",
            "Остаточная трудоемкость, ч",
        ],
        data=detailed_orders_data,
        column_widths={"A": 22, "B": 44, "C": 44, "D": 22, "E": 22, "F": 22},
        style_columns=["C", "D", "E", "F"],
        filter_columns=["A", "B", "C"],
    )

    file: BytesIO = BytesIO()
    workbook.save(file)
    file.seek(0)
    return file
